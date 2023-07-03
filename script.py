# import module
import openpyxl
import requests

API_ENDPOINT = "http://ec2-15-207-71-215.ap-south-1.compute.amazonaws.com:4000/patents/patent"

# load excel with its path
wrkbk = openpyxl.load_workbook("./Patentstracking.xlsx")

sh = wrkbk.active

# iterate through excel and display data
for i in range(2, 106):
    data = {
        "Reference" : sh.cell(row=i, column=1).value,
        "Title": sh.cell(row=i, column=2).value,
        "Inventor_List": sh.cell(row=i, column=3).value,
        "Center_Name": sh.cell(row=i, column=4).value,
        "Patent_Number": sh.cell(row=i, column=5).value,
        "ProvisionalFilingDate": sh.cell(row=i, column=6).value,
        "FullFilingDate": sh.cell(row=i, column=7).value,
        "YearofProvisionalFiling": sh.cell(row=i, column=8).value,
        "YearofFullFiling": sh.cell(row=i, column=9).value,
        "Year": sh.cell(row=i, column=10).value,
        "YearofGrant": sh.cell(row=i, column=11).value,
        "FilledinCountry": sh.cell(row=i, column=12).value,
        "PatentStatusComment": sh.cell(row=i, column=13).value,
        "FacultyName": sh.cell(row=i, column=15).value,
        "Funding": sh.cell(row=i, column=16).value,
        "CollaboratorsEmailId": sh.cell(row=i, column=17).value,
        "NoStudentInventors": sh.cell(row=i, column=18).value,
        "IDFSubmissionDate": sh.cell(row=i, column=19).value,
        "IDFSerachReportDate": sh.cell(row=i, column=20).value,
        "ApprovalDateforApplication": sh.cell(row=i, column=21).value,
        "FirstDraft": sh.cell(row=i, column=22).value,
        "ProvisionalCompleteDraftdate": sh.cell(row=i, column=23).value,
        "TotalCost": sh.cell(row=i, column=24).value,
        "Comments": sh.cell(row=i, column=25).value,
    }
    # print(data)
    r = requests.post(url=API_ENDPOINT, data=data)
    pastebin_url = r.text
    print("The pastebin URL is:%s"%pastebin_url)
